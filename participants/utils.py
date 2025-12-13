import io
import random
from typing import Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


DEV_KEYWORDS = {
    "DEVELOPPEMENT BACKEND",
    "DEVELOPPEMENT FRONTEND",
    "DEVELOPPEMENT FULLSTACK",
    "MODELISATION DES SYSTEMES D'INFORMATION",
    "SECURITE RESEAUX",
}

MARKETING_KEYWORDS = {
    "COMMUNITY MANAGEMENT",
    "MEDIA BUYER",
    "STORYTELLING",
    "COPYWRITING",
}

LANG_TOKENS_FR = ("fr", "fra", "fran", "franc", "french", "francais")
LANG_TOKENS_EN = ("en", "ang", "eng", "anglais", "english")

ACADEMIC_SCORES = {"B1": 1, "B2": 2, "B3": 3, "M1": 4, "M2": 5}

USEFUL_COLUMNS = [
    "NOM ET PRENOM",
    "Email Address",
    "LANGUE",
    "NIVEAU D'ETUDES",
    "VOS COMPETENCES",
]


def _clean_text(value) -> str:
    if pd.isna(value):
        return ""
    return str(value).strip()


def _split_skills(raw_value: str) -> List[str]:
    cleaned = _clean_text(raw_value)
    if not cleaned:
        return []
    separators = [",", ";", "/", "|", "\n"]
    for sep in separators:
        cleaned = cleaned.replace(sep, ",")
    return [item.strip().upper() for item in cleaned.split(",") if item.strip()]


def parse_participants(uploaded_file) -> Tuple[List[Dict], List[str]]:
    """Read the Excel file and return structured participants plus the column order."""
    frame = pd.read_excel(uploaded_file)
    frame = frame.fillna("")
    columns = [col for col in USEFUL_COLUMNS if col in frame.columns]
    if not columns:
        columns = USEFUL_COLUMNS

    participants = []
    for idx, (_, row) in enumerate(frame.iterrows()):
        record = {col: _clean_text(row[col]) for col in columns if col in frame.columns}
        record["uid"] = f"p{idx}"
        participants.append(_enrich_participant(record))

    return participants, columns


def _enrich_participant(record: Dict) -> Dict:
    language_raw = _clean_text(record.get("LANGUE", record.get("Langue", "")))
    language_lower = language_raw.lower()
    skills = _split_skills(record.get("VOS COMPETENCES", record.get("Competences", "")))
    level_raw = _clean_text(record.get("NIVEAU D'ETUDES", record.get("Niveau d'etudes", ""))).upper()

    language_fr = any(token in language_lower for token in LANG_TOKENS_FR) or "les deux" in language_lower
    language_en = any(token in language_lower for token in LANG_TOKENS_EN) or "les deux" in language_lower

    is_dev = any(skill in DEV_KEYWORDS for skill in skills)
    is_marketing = any(skill in MARKETING_KEYWORDS for skill in skills)

    enriched = {
        **record,
        "uid": record.get("uid"),
        "skills_list": skills,
        "language_raw": language_raw or "Non precise",
        "language_fr": language_fr,
        "language_en": language_en,
        "is_dev": is_dev,
        "is_marketing": is_marketing,
        "academic_level": level_raw or "NC",
        "academic_score": ACADEMIC_SCORES.get(level_raw, 0),
        "email_sent": record.get("email_sent", False),
        "team": record.get("team", None),
        "team_display": record.get("team_display", None),
        "is_leader": record.get("is_leader", False),
    }
    return enriched


def build_email_content(language: str, full_name: str) -> Tuple[str, str]:
    name = full_name or "participant"
    english = (
        f"Hello {name},\n\n"
        "Welcome to the EEUEZ hackathon! You are registered. "
        "We will share your team assignment and workshop details soon.\n\n"
        "Stay tuned.\nTeam EEUEZ"
    )
    french = (
        f"Bonjour {name},\n\n"
        "Bienvenue au hackathon EEUEZ ! Votre inscription est bien prise en compte. "
        "Nous vous communiquerons bientot votre equipe et les informations des ateliers.\n\n"
        "A tres vite,\nEquipe EEUEZ"
    )

    lang_lower = (language or "").lower()
    if "les deux" in lang_lower:
        return ("EEUEZ Hackathon / Hackathon EEUEZ", f"{french}\n\n----\n\n{english}")
    if any(token in lang_lower for token in LANG_TOKENS_FR):
        return ("Hackathon EEUEZ - Confirmation", french)
    return ("EEUEZ Hackathon - Confirmation", english)


def assign_teams(participants: List[Dict], team_names: Optional[Dict[str, str]] = None) -> List[Dict]:
    """Assign participants into up to 10 teams of 5 with soft constraints."""
    for person in participants:
        person["team"] = None
        person["team_display"] = None
        person["is_leader"] = False

    remaining = participants.copy()
    random.shuffle(remaining)
    teams: List[Dict] = []

    def pop_first(predicate):
        for idx, candidate in enumerate(remaining):
            if predicate(candidate):
                return remaining.pop(idx)
        return None

    for idx in range(10):
        if not remaining:
            break

        team_name = f"TEAM {idx + 1}"
        display_name = (team_names or {}).get(team_name) or team_name
        team_members: List[Dict] = []

        def ensure(predicate):
            if len(team_members) >= 5:
                return
            if any(predicate(member) for member in team_members):
                return
            picked = pop_first(predicate)
            if picked:
                team_members.append(picked)

        ensure(lambda p: p.get("is_dev"))
        ensure(lambda p: p.get("is_marketing"))
        ensure(lambda p: p.get("language_en"))
        ensure(lambda p: p.get("language_fr"))

        while len(team_members) < 5 and remaining:
            team_members.append(remaining.pop(0))

        leader = _pick_leader(team_members)
        for member in team_members:
            member["team"] = team_name
            member["team_display"] = display_name
            member["is_leader"] = leader is not None and member is leader

        teams.append({"name": team_name, "display_name": display_name, "members": team_members, "leader": leader})

    return teams


def _pick_leader(team_members: List[Dict]):
    if not team_members:
        return None
    sorted_members = sorted(team_members, key=lambda m: m.get("academic_score", 0), reverse=True)
    return sorted_members[0]


def build_report_workbook(
    participants: List[Dict],
    teams: List[Dict],
    source_columns: Optional[List[str]] = None,
) -> bytes:
    """Create the Excel report with the general sheet plus 10 team sheets."""
    wb = Workbook()
    general_ws = wb.active
    general_ws.title = "General"
    header_font = Font(bold=True)

    base_headers = USEFUL_COLUMNS
    columns = source_columns or base_headers
    general_headers = columns + ["Team", "Role", "Email envoye"]
    general_ws.append(general_headers)
    for cell in general_ws[1]:
        cell.font = header_font

    for person in participants:
        row = [person.get(col, "") for col in columns]
        row.extend(
            [
                person.get("team_display") or person.get("team") or "Non assigne",
                "Chef d'equipe" if person.get("is_leader") else "Membre",
                "Oui" if person.get("email_sent") else "Non",
            ]
        )
        general_ws.append(row)

    for team in teams:
        _build_team_sheet(
            workbook=wb,
            sheet_name=team["name"],
            display_name=team.get("display_name") or team["name"],
            mentor=team.get("mentor"),
            members=team.get("members", []),
            header_font=header_font,
        )

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def _build_team_sheet(
    workbook: Workbook,
    sheet_name: str,
    display_name: str,
    mentor: Optional[Dict],
    members: List[Dict],
    header_font: Font,
):
    ws = workbook.create_sheet(sheet_name)
    ws["A1"] = "Nom de l'equipe / Team name:"
    ws["B1"] = display_name
    ws["A2"] = "Encadrant"
    if mentor:
        mentor_line = mentor.get("name") or ""
        if mentor.get("email"):
            mentor_line = f"{mentor_line} ({mentor.get('email')})"
        ws["B2"] = mentor_line
    else:
        ws["B2"] = "Non assigne"
    ws.append([])  # Blank line

    headers = ["Nom complet", "Email", "Langue", "Niveau", "Competences", "Role"]
    for idx in range(1, 9):
        headers.append(f"Atelier {idx}")
    headers.append("Total (/20)")

    ws.append(headers)
    for cell in ws[4]:
        cell.font = header_font

    for row_idx, member in enumerate(members, start=5):
        row_values = [
            member.get("NOM ET PRENOM") or member.get("Nom") or "",
            member.get("Email Address") or member.get("EMAIL") or "",
            member.get("language_raw", member.get("LANGUE", "")),
            member.get("academic_level", member.get("NIVEAU D'ETUDES", "")),
            ", ".join(member.get("skills_list", [])) or member.get("VOS COMPETENCES", ""),
            "Chef d'equipe" if member.get("is_leader") else "Membre",
        ]
        # Empty placeholders for atelier scores
        row_values.extend([""] * 8)
        row_values.append("")  # placeholder for total formula
        ws.append(row_values)

        total_col = get_column_letter(6 + 8 + 1)  # after role + eight ateliers
        first_atelier_col = get_column_letter(7)
        last_atelier_col = get_column_letter(14)
        total_cell = f"{total_col}{row_idx}"
        ws[total_cell] = (
            f'=IF(COUNT({first_atelier_col}{row_idx}:{last_atelier_col}{row_idx})=0,"",'
            f'ROUND(AVERAGE({first_atelier_col}{row_idx}:{last_atelier_col}{row_idx}),2))'
        )
